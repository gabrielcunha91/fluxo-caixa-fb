import streamlit as st
import pandas as pd
import openpyxl
import os

st.set_page_config(
    page_title="Conciliacao",
    page_icon="📃",
    layout="wide"
)

df_lojas = st.session_state["lojas"]

lojas = df_lojas["Loja"].unique()
loja = st.selectbox("Loja", lojas)

# Defina um dicionário para mapear nomes de lojas a IDs de lojas
mapeamento_lojas = dict(zip(df_lojas["Loja"], df_lojas["ID_Loja"]))

# Obtenha o ID da loja selecionada
id_loja = mapeamento_lojas[loja]

st.write(id_loja)

# def export_to_excel(df, sheet_name, excel_filename):
#     if os.path.exists(excel_filename):
#         wb = openpyxl.load_workbook(excel_filename)
#     else:
#         wb = openpyxl.Workbook()

#     if sheet_name in wb.sheetnames:
#         wb.remove(wb[sheet_name])

#     ws = wb.create_sheet(title=sheet_name)
#     for r_idx, row in enumerate(df.iterrows(), start=1):
#         for c_idx, value in enumerate(row[1], start=1):
#             ws.cell(row=r_idx, column=c_idx, value=value)

#     wb.save(excel_filename)

def export_to_excel(df, sheet_name, excel_filename):
    if os.path.exists(excel_filename):
        wb = openpyxl.load_workbook(excel_filename)
    else:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)
    
    # Escrever os cabeçalhos
    for col_idx, column_title in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=column_title)
    
    # Escrever os dados
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(excel_filename)


# Definindo o nome do arquivo excel
excel_filename = 'Conciliacao_FB.xlsx'

st.markdown('---')
st.markdown("Faturamento Zig")

df_faturam_zig = st.session_state["faturam_zig"]
df_faturam_zig_loja = df_faturam_zig[df_faturam_zig['ID_Loja'] == id_loja]
df_faturam_zig_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Faturam Zig'):
    sheet_name_zig = 'df_faturam_zig'
    export_to_excel(df_faturam_zig_loja, sheet_name_zig, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("Receitas Extraordinárias")

df_receitas_extraord = st.session_state["receitas_extraord"]
df_receitas_extraord_loja = df_receitas_extraord[df_receitas_extraord['ID_Loja'] == id_loja]
df_receitas_extraord_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Receitas Extraord'):
    sheet_name_receitas_extraord = 'df_receitas_extraord'
    export_to_excel(df_receitas_extraord_loja, sheet_name_receitas_extraord, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("View Parcelamentos Agrupados - Receitas Extraord")

df_view_parc_agrup = st.session_state["view_parc_agrup"]
df_view_parc_loja = df_view_parc_agrup[df_view_parc_agrup['ID_Loja'] == id_loja]
df_view_parc_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar View Parcelamentos Receitas Extraord'):
    sheet_name_view_parc_agrup = 'view_parc_agrup'
    export_to_excel(df_view_parc_loja, sheet_name_view_parc_agrup, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("Custos BlueMe Sem Parcelamento")

df_custos_blueme_sem_parcelamento = st.session_state["custos_blueme_sem_parcelamento"]
df_custos_blueme_sem_parcelamento_loja = df_custos_blueme_sem_parcelamento[df_custos_blueme_sem_parcelamento['ID_Loja'] == id_loja]
df_custos_blueme_sem_parcelamento_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Custos BlueMe Sem Parcelamento'):
    sheet_name_custos_blueme_sem_parcelamento = 'df_blueme_sem_parcelamento'
    export_to_excel(df_custos_blueme_sem_parcelamento_loja, sheet_name_custos_blueme_sem_parcelamento, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("Custos BlueMe Com Parcelamento")

df_custos_blueme_com_parcelamento = st.session_state["custos_blueme_com_parcelamento"]
df_custos_blueme_com_parcelamento_loja = df_custos_blueme_com_parcelamento[df_custos_blueme_com_parcelamento['ID_Loja'] == id_loja]
df_custos_blueme_com_parcelamento_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Custos BlueMe Com Parcelamento'):
    sheet_name_custos_blueme_com_parcelamento = 'df_blueme_com_parcelamento'
    export_to_excel(df_custos_blueme_com_parcelamento_loja, sheet_name_custos_blueme_com_parcelamento, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("Extratos Bancários")

df_extratos = st.session_state["extratos_bancarios"]
df_extratos_loja = df_extratos[df_extratos['ID_Loja'] == id_loja]
df_extratos_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Extratos'):
    sheet_name_extratos = 'df_extratos'
    export_to_excel(df_extratos_loja, sheet_name_extratos, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("Mutuos")

df_mutuos = st.session_state["mutuos"]
df_mutuos_loja = df_mutuos[((df_mutuos['ID_Loja_Saida'] == id_loja) | (df_mutuos['ID_Loja_Entrada'] == id_loja))]
df_mutuos_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Mutuos'):
    df_mutuos_loja['Valor_Entrada'] = df_mutuos_loja.apply(lambda row: row['Valor'] if row['ID_Loja_Entrada'] == id_loja else 0, axis=1)
    df_mutuos_loja['Valor_Saida'] = df_mutuos_loja.apply(lambda row: row['Valor'] if row['ID_Loja_Saida'] == id_loja else 0, axis=1)
    df_mutuos_loja = df_mutuos_loja.drop('Valor', axis=1)

    sheet_name_mutuos = 'df_mutuos'
    export_to_excel(df_mutuos_loja, sheet_name_mutuos, excel_filename)
    st.success('Arquivo atualizado com sucesso!')

st.markdown('---')
st.markdown("Tesouraria - Transações")

df_tesouraria_trans = st.session_state["tesouraria_trans"]
df_tesouraria_trans_loja = df_tesouraria_trans[df_tesouraria_trans['ID_Loja'] == id_loja]
df_tesouraria_trans_loja
# Chama a função para atualizar o arquivo Excel
if st.button('Atualizar Tesouraria Transações'):
    sheet_name_tesouraria = 'df_tesouraria_trans'
    export_to_excel(df_tesouraria_trans_loja, sheet_name_tesouraria, excel_filename)
    st.success('Arquivo atualizado com sucesso!')



st.markdown('---')

if st.button('Baixar Excel'):
    if os.path.exists(excel_filename):
        with open(excel_filename, "rb") as file:
            file_content = file.read()
        st.download_button(
            label="Clique para baixar o arquivo Excel",
            data=file_content,
            file_name="Conciliacao_FB.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )



